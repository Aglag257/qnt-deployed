import streamlit as st
import re
from io import BytesIO
from openpyxl import Workbook

st.title("Text to Excel Converter")

user_input = st.text_area("Paste your data below:", height=300)

if st.button("Convert to Excel"):
    if user_input.strip():
        wb = Workbook()
        ws = wb.active

        lines = user_input.strip().split("\n")
        skipped_lines = []

        for row_num, line in enumerate(lines, start=1):

            parts = re.split(r'\s+', line.strip())
            if len(parts) >= 2:
                ws.cell(row=row_num, column=1).value = parts[0].replace(",", ".")
                ws.cell(row=row_num, column=2).value = parts[1].replace(",", ".")
            else:
                skipped_lines.append((row_num, line.strip()))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Excel file created!")

        st.download_button(
            label="Download Excel file",
            data=output,
            file_name="output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if skipped_lines:
            st.warning("Some lines were skipped:")
            for row_num, content in skipped_lines:
                st.text(f"Line {row_num}: {content}")
    else:
        st.error("Please paste some data first.")
